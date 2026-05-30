"""
=======================================================================
OLIST PROJECT — Python Automation: SQL → Excel Export
=======================================================================
File:    python/exports/02_export_to_excel.py
Purpose: Execute all analytics SQL queries via Python, transform
         results, and export a professional multi-sheet Excel report.

Run after: 01_db_setup.py (database must be populated)
Output:    outputs/olist_analytics_report.xlsx
=======================================================================
"""

import sys
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

# Import our connection utility
sys.path.append(str(Path(__file__).parent.parent / "connection"))
from db_setup import get_engine, run_query

OUTPUT_DIR = Path("../../outputs")
OUTPUT_DIR.mkdir(exist_ok=True)

REPORT_FILE = OUTPUT_DIR / f"olist_analytics_{datetime.now().strftime('%Y%m%d')}.xlsx"


# -----------------------------------------------------------------------
# SQL Queries — each becomes one Excel sheet
# -----------------------------------------------------------------------

QUERIES = {
    "1_Monthly_GMV": """
        SELECT
            DATE_TRUNC('month', o.order_purchase_timestamp)::DATE AS order_month,
            COUNT(DISTINCT o.order_id)                            AS total_orders,
            ROUND(SUM(p.payment_value)::NUMERIC, 2)              AS gmv,
            ROUND(AVG(p.payment_value)::NUMERIC, 2)              AS avg_order_value,
            ROUND(SUM(i.freight_value)::NUMERIC, 2)              AS total_freight
        FROM olist.olist_orders o
        JOIN olist.olist_order_payments p ON o.order_id = p.order_id
        JOIN olist.olist_order_items    i ON o.order_id = i.order_id
        WHERE o.order_status NOT IN ('canceled','unavailable')
          AND o.order_purchase_timestamp IS NOT NULL
        GROUP BY 1
        ORDER BY 1
    """,

    "2_Category_Revenue": """
        SELECT
            COALESCE(t.product_category_name_english, p.product_category_name, 'Unknown') AS category,
            COUNT(DISTINCT o.order_id)   AS total_orders,
            ROUND(SUM(oi.price)::NUMERIC, 2)         AS gross_revenue,
            ROUND(AVG(oi.price)::NUMERIC, 2)         AS avg_price,
            ROUND(AVG(r.review_score)::NUMERIC, 2)   AS avg_review_score
        FROM olist.olist_order_items oi
        JOIN olist.olist_orders       o  ON oi.order_id  = o.order_id
        JOIN olist.olist_products     p  ON oi.product_id = p.product_id
        LEFT JOIN olist.product_category_name_translation t
                                         ON p.product_category_name = t.product_category_name
        LEFT JOIN olist.olist_order_reviews r ON o.order_id = r.order_id
        WHERE o.order_status NOT IN ('canceled','unavailable')
        GROUP BY 1
        ORDER BY gross_revenue DESC
        LIMIT 30
    """,

    "3_Delivery_By_State": """
        SELECT
            c.customer_state,
            COUNT(*)                                                      AS total_delivered,
            ROUND(AVG(EXTRACT(EPOCH FROM (o.order_delivered_customer_date
                - o.order_purchase_timestamp)) / 86400.0)::NUMERIC, 1)   AS avg_delivery_days,
            ROUND(AVG(EXTRACT(EPOCH FROM (o.order_delivered_customer_date
                - o.order_estimated_delivery_date)) / 86400.0)::NUMERIC, 1) AS avg_delay_days,
            ROUND(
                SUM(CASE WHEN o.order_delivered_customer_date
                    <= o.order_estimated_delivery_date THEN 1 ELSE 0 END)::NUMERIC
                / COUNT(*) * 100, 1
            )                                                             AS on_time_rate_pct
        FROM olist.olist_orders    o
        JOIN olist.olist_customers c ON o.customer_id = c.customer_id
        WHERE o.order_status = 'delivered'
          AND o.order_delivered_customer_date IS NOT NULL
          AND o.order_estimated_delivery_date IS NOT NULL
        GROUP BY c.customer_state
        ORDER BY avg_delay_days DESC
    """,

    "4_Review_vs_Delay": """
        SELECT
            CASE
                WHEN o.order_delivered_customer_date <= o.order_estimated_delivery_date
                    THEN 'On time or early'
                WHEN EXTRACT(EPOCH FROM (o.order_delivered_customer_date
                    - o.order_estimated_delivery_date)) / 86400 <= 3
                    THEN 'Slightly late (1–3 days)'
                WHEN EXTRACT(EPOCH FROM (o.order_delivered_customer_date
                    - o.order_estimated_delivery_date)) / 86400 <= 7
                    THEN 'Late (4–7 days)'
                ELSE 'Very late (>7 days)'
            END                                      AS delivery_bucket,
            COUNT(*)                                 AS order_count,
            ROUND(AVG(r.review_score)::NUMERIC, 3)   AS avg_review_score,
            ROUND(SUM(CASE WHEN r.review_score = 1 THEN 1 ELSE 0 END)::NUMERIC
                / COUNT(*) * 100, 1)                 AS pct_1_star,
            ROUND(SUM(CASE WHEN r.review_score = 5 THEN 1 ELSE 0 END)::NUMERIC
                / COUNT(*) * 100, 1)                 AS pct_5_star
        FROM olist.olist_orders         o
        JOIN olist.olist_order_reviews  r ON o.order_id = r.order_id
        WHERE o.order_status = 'delivered'
          AND o.order_delivered_customer_date IS NOT NULL
        GROUP BY 1
        ORDER BY avg_review_score DESC
    """,

    "5_Seller_Performance": """
        SELECT
            s.seller_id,
            s.seller_state,
            COUNT(DISTINCT oi.order_id)              AS total_orders,
            ROUND(SUM(oi.price)::NUMERIC, 2)         AS gross_revenue,
            ROUND(AVG(oi.price)::NUMERIC, 2)         AS avg_item_price,
            ROUND(AVG(r.review_score)::NUMERIC, 3)   AS avg_review_score,
            ROUND(AVG(EXTRACT(EPOCH FROM (
                o.order_delivered_carrier_date - o.order_purchase_timestamp
            )) / 86400.0)::NUMERIC, 1)               AS avg_fulfillment_days
        FROM olist.olist_sellers      s
        JOIN olist.olist_order_items oi ON s.seller_id = oi.seller_id
        JOIN olist.olist_orders       o  ON oi.order_id = o.order_id
        LEFT JOIN olist.olist_order_reviews r ON o.order_id = r.order_id
        WHERE o.order_status = 'delivered'
        GROUP BY s.seller_id, s.seller_state
        HAVING COUNT(DISTINCT oi.order_id) >= 10
        ORDER BY gross_revenue DESC
        LIMIT 50
    """,

    "6_RFM_Segments": """
        WITH rfm AS (
            SELECT
                c.customer_unique_id,
                DATE_PART('day', DATE '2018-10-17' - MAX(o.order_purchase_timestamp)::DATE)
                                              AS recency_days,
                COUNT(DISTINCT o.order_id)    AS frequency,
                ROUND(SUM(p.payment_value)::NUMERIC, 2) AS monetary
            FROM olist.olist_customers       c
            JOIN olist.olist_orders          o ON c.customer_id = o.customer_id
            JOIN olist.olist_order_payments  p ON o.order_id   = p.order_id
            WHERE o.order_status NOT IN ('canceled','unavailable')
            GROUP BY c.customer_unique_id
        ),
        scored AS (
            SELECT *,
                NTILE(5) OVER (ORDER BY recency_days ASC) AS r_score,
                NTILE(5) OVER (ORDER BY frequency DESC)   AS f_score,
                NTILE(5) OVER (ORDER BY monetary DESC)    AS m_score
            FROM rfm
        )
        SELECT
            CASE
                WHEN r_score >= 4 AND f_score >= 4 AND m_score >= 4 THEN 'Champions'
                WHEN r_score >= 3 AND f_score >= 3                  THEN 'Loyal Customers'
                WHEN r_score >= 4 AND f_score <= 2                  THEN 'Recent Customers'
                WHEN r_score <= 2 AND f_score >= 3 AND m_score >= 3 THEN 'At Risk'
                WHEN r_score <= 2 AND f_score <= 2                  THEN 'Churned'
                ELSE 'Potential Loyalists'
            END                          AS rfm_segment,
            COUNT(*)                     AS customer_count,
            ROUND(AVG(recency_days), 0)  AS avg_recency_days,
            ROUND(AVG(frequency), 2)     AS avg_frequency,
            ROUND(AVG(monetary), 2)      AS avg_monetary
        FROM scored
        GROUP BY 1
        ORDER BY avg_monetary DESC
    """,
}


# -----------------------------------------------------------------------
# Excel Formatting Helpers
# -----------------------------------------------------------------------

HEADER_FILL  = PatternFill("solid", fgColor="1A3C5E")   # Dark navy
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL     = PatternFill("solid", fgColor="F0F4F8")   # Light blue-grey
BORDER_THIN  = Border(
    bottom=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin",  color="EEEEEE"),
)


def style_sheet(ws):
    """Apply professional formatting to a worksheet."""
    # Header row
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER_THIN

    # Data rows — alternate shading
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for cell in row:
            if fill:
                cell.fill = fill
            cell.border    = BORDER_THIN
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # Freeze header row
    ws.freeze_panes = "A2"


# -----------------------------------------------------------------------
# Main Export Function
# -----------------------------------------------------------------------

def export_to_excel():
    print("🔌 Connecting to PostgreSQL...")
    engine = get_engine()

    print(f"📊 Running {len(QUERIES)} queries and exporting to Excel...")
    print(f"   Output: {REPORT_FILE}")

    with pd.ExcelWriter(REPORT_FILE, engine="openpyxl") as writer:
        for sheet_name, sql in QUERIES.items():
            print(f"   → {sheet_name}...", end=" ")
            try:
                df = run_query(sql, engine)
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                print(f"✅ ({len(df):,} rows)")
            except Exception as e:
                print(f"❌ Error: {e}")

        # Add a metadata sheet
        meta = pd.DataFrame({
            "Field": ["Project", "Dataset", "Generated", "Total Queries", "Author"],
            "Value": [
                "Olist E-Commerce Analysis",
                "Brazilian E-Commerce Public Dataset (Kaggle)",
                datetime.now().strftime("%Y-%m-%d %H:%M"),
                len(QUERIES),
                "Your Name Here",
            ]
        })
        meta.to_excel(writer, sheet_name="README", index=False)

    # Apply styling to all sheets
    print("\n🎨 Applying professional formatting...")
    wb = load_workbook(REPORT_FILE)
    for sheet in wb.sheetnames:
        style_sheet(wb[sheet])
    wb.save(REPORT_FILE)

    print(f"\n✅ Export complete: {REPORT_FILE}")
    return REPORT_FILE


if __name__ == "__main__":
    export_to_excel()
