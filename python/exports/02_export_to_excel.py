"""
OLIST PROJECT - Export SQL results to Excel
Platform: macOS (python3 / pip3)
Run: python3 02_export_to_excel.py
"""

import os
import time
import pandas as pd
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv
from sqlalchemy import create_engine, text
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Load credentials from .env ---
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
ENV_FILE     = PROJECT_ROOT / ".env"
load_dotenv(dotenv_path=ENV_FILE)

DB_USER     = os.getenv("DB_USER", "")
DB_PASSWORD = os.getenv("DB_PASSWORD", "")
DB_HOST     = os.getenv("DB_HOST", "localhost")
DB_PORT     = os.getenv("DB_PORT", "5432")
DB_NAME     = os.getenv("DB_NAME", "olist_db")

CONN = f"postgresql+psycopg2://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
OUTPUT_DIR = PROJECT_ROOT / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)
REPORT_FILE = OUTPUT_DIR / f"olist_analytics_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

def get_engine():
    return create_engine(
        CONN,
        pool_pre_ping=True,
        connect_args={"options": "-csearch_path=olist"}
    )

QUERIES = {
    "1_Monthly_GMV": """
        SELECT
            DATE_TRUNC('month', o.order_purchase_timestamp)::DATE AS order_month,
            COUNT(DISTINCT o.order_id)                            AS total_orders,
            ROUND(SUM(p.payment_value)::NUMERIC, 2)              AS gmv,
            ROUND(AVG(p.payment_value)::NUMERIC, 2)              AS avg_order_value,
            ROUND(SUM(i.freight_value)::NUMERIC, 2)              AS total_freight
        FROM olist.olist_orders o
        JOIN olist.olist_order_payments p ON o.order_id = p.order_id
        JOIN olist.olist_order_items    i ON o.order_id = i.order_id
        WHERE o.order_status NOT IN ('canceled','unavailable')
          AND o.order_purchase_timestamp IS NOT NULL
        GROUP BY 1 ORDER BY 1
    """,
    "2_Category_Revenue": """
        SELECT
            COALESCE(t.product_category_name_english,
                     p.product_category_name, 'Unknown')  AS category,
            COUNT(DISTINCT o.order_id)                    AS total_orders,
            ROUND(SUM(oi.price)::NUMERIC, 2)              AS gross_revenue,
            ROUND(AVG(oi.price)::NUMERIC, 2)              AS avg_price,
            ROUND(AVG(r.review_score)::NUMERIC, 2)        AS avg_review_score
        FROM olist.olist_order_items oi
        JOIN olist.olist_orders       o  ON oi.order_id   = o.order_id
        JOIN olist.olist_products     p  ON oi.product_id = p.product_id
        LEFT JOIN olist.product_category_name_translation t
                                          ON p.product_category_name = t.product_category_name
        LEFT JOIN olist.olist_order_reviews r ON o.order_id = r.order_id
        WHERE o.order_status NOT IN ('canceled','unavailable')
        GROUP BY 1 ORDER BY gross_revenue DESC LIMIT 30
    """,
    "3_Delivery_By_State": """
        SELECT
            c.customer_state,
            COUNT(*)                                                        AS total_delivered,
            ROUND(AVG(EXTRACT(EPOCH FROM (o.order_delivered_customer_date
                - o.order_purchase_timestamp)) / 86400.0)::NUMERIC, 1)     AS avg_delivery_days,
            ROUND(AVG(EXTRACT(EPOCH FROM (o.order_delivered_customer_date
                - o.order_estimated_delivery_date)) / 86400.0)::NUMERIC, 1) AS avg_delay_days,
            ROUND(
                SUM(CASE WHEN o.order_delivered_customer_date
                    <= o.order_estimated_delivery_date THEN 1 ELSE 0 END)::NUMERIC
                / COUNT(*) * 100, 1
            )                                                               AS on_time_rate_pct
        FROM olist.olist_orders    o
        JOIN olist.olist_customers c ON o.customer_id = c.customer_id
        WHERE o.order_status = 'delivered'
          AND o.order_delivered_customer_date IS NOT NULL
          AND o.order_estimated_delivery_date IS NOT NULL
        GROUP BY c.customer_state
        ORDER BY avg_delay_days DESC
    """,
    "4_Review_vs_Delay": """
        SELECT
            CASE
                WHEN o.order_delivered_customer_date <= o.order_estimated_delivery_date
                    THEN '1. On time or early'
                WHEN EXTRACT(EPOCH FROM (o.order_delivered_customer_date
                    - o.order_estimated_delivery_date)) / 86400 <= 3
                    THEN '2. Slightly late (1-3 days)'
                WHEN EXTRACT(EPOCH FROM (o.order_delivered_customer_date
                    - o.order_estimated_delivery_date)) / 86400 <= 7
                    THEN '3. Late (4-7 days)'
                ELSE '4. Very late (>7 days)'
            END                                       AS delivery_bucket,
            COUNT(*)                                  AS order_count,
            ROUND(AVG(r.review_score)::NUMERIC, 3)    AS avg_review_score,
            ROUND(SUM(CASE WHEN r.review_score = 1 THEN 1 ELSE 0 END)::NUMERIC
                / COUNT(*) * 100, 1)                  AS pct_1_star,
            ROUND(SUM(CASE WHEN r.review_score = 5 THEN 1 ELSE 0 END)::NUMERIC
                / COUNT(*) * 100, 1)                  AS pct_5_star
        FROM olist.olist_orders         o
        JOIN olist.olist_order_reviews  r ON o.order_id = r.order_id
        WHERE o.order_status = 'delivered'
          AND o.order_delivered_customer_date IS NOT NULL
        GROUP BY 1 ORDER BY 1
    """,
    "5_Seller_Performance": """
        SELECT
            s.seller_id,
            s.seller_state,
            COUNT(DISTINCT oi.order_id)              AS total_orders,
            ROUND(SUM(oi.price)::NUMERIC, 2)         AS gross_revenue,
            ROUND(AVG(oi.price)::NUMERIC, 2)         AS avg_item_price,
            ROUND(AVG(r.review_score)::NUMERIC, 3)   AS avg_review_score,
            ROUND(AVG(EXTRACT(EPOCH FROM (
                o.order_delivered_carrier_date - o.order_purchase_timestamp
            )) / 86400.0)::NUMERIC, 1)               AS avg_fulfillment_days
        FROM olist.olist_sellers      s
        JOIN olist.olist_order_items oi ON s.seller_id = oi.seller_id
        JOIN olist.olist_orders       o  ON oi.order_id = o.order_id
        LEFT JOIN olist.olist_order_reviews r ON o.order_id = r.order_id
        WHERE o.order_status = 'delivered'
        GROUP BY s.seller_id, s.seller_state
        HAVING COUNT(DISTINCT oi.order_id) >= 10
        ORDER BY gross_revenue DESC LIMIT 50
    """,
    "6_RFM_Segments": """
        SELECT
            rfm_segment,
            COUNT(*)                             AS customer_count,
            ROUND(AVG(recency_days)::NUMERIC, 0) AS avg_recency_days,
            ROUND(AVG(frequency)::NUMERIC, 2)    AS avg_frequency,
            ROUND(AVG(monetary)::NUMERIC, 2)     AS avg_monetary
        FROM (
            SELECT
                c.customer_unique_id,
                (DATE '2018-10-17' - MAX(o.order_purchase_timestamp)::DATE) AS recency_days,
                COUNT(DISTINCT o.order_id)   AS frequency,
                SUM(p.payment_value)         AS monetary,
                CASE
                    WHEN NTILE(5) OVER (ORDER BY MAX(o.order_purchase_timestamp) DESC) >= 4
                     AND NTILE(5) OVER (ORDER BY COUNT(DISTINCT o.order_id) DESC) >= 4
                     AND NTILE(5) OVER (ORDER BY SUM(p.payment_value) DESC) >= 4
                        THEN 'Champions'
                    WHEN NTILE(5) OVER (ORDER BY MAX(o.order_purchase_timestamp) DESC) >= 3
                     AND NTILE(5) OVER (ORDER BY COUNT(DISTINCT o.order_id) DESC) >= 3
                        THEN 'Loyal Customers'
                    WHEN NTILE(5) OVER (ORDER BY MAX(o.order_purchase_timestamp) DESC) >= 4
                     AND NTILE(5) OVER (ORDER BY COUNT(DISTINCT o.order_id) DESC) <= 2
                        THEN 'Recent Customers'
                    WHEN NTILE(5) OVER (ORDER BY MAX(o.order_purchase_timestamp) DESC) <= 2
                     AND NTILE(5) OVER (ORDER BY COUNT(DISTINCT o.order_id) DESC) >= 3
                     AND NTILE(5) OVER (ORDER BY SUM(p.payment_value) DESC) >= 3
                        THEN 'At Risk'
                    WHEN NTILE(5) OVER (ORDER BY MAX(o.order_purchase_timestamp) DESC) <= 2
                     AND NTILE(5) OVER (ORDER BY COUNT(DISTINCT o.order_id) DESC) <= 2
                        THEN 'Churned'
                    ELSE 'Potential Loyalists'
                END AS rfm_segment
            FROM olist.olist_customers      c
            JOIN olist.olist_orders         o ON c.customer_id = o.customer_id
            JOIN olist.olist_order_payments p ON o.order_id    = p.order_id
            WHERE o.order_status NOT IN ('canceled','unavailable')
            GROUP BY c.customer_unique_id
        ) scored
        GROUP BY rfm_segment
        ORDER BY avg_monetary DESC
    """,
    "7_Customer_Retention": """
        WITH counts AS (
            SELECT
                c.customer_unique_id,
                COUNT(DISTINCT o.order_id) AS order_count
            FROM olist.olist_customers c
            JOIN olist.olist_orders    o ON c.customer_id = o.customer_id
            WHERE o.order_status NOT IN ('canceled','unavailable')
            GROUP BY c.customer_unique_id
        )
        SELECT
            COUNT(*)                                                   AS total_customers,
            SUM(CASE WHEN order_count = 1  THEN 1 ELSE 0 END)         AS one_time_buyers,
            SUM(CASE WHEN order_count >= 2 THEN 1 ELSE 0 END)         AS repeat_buyers,
            ROUND(SUM(CASE WHEN order_count >= 2 THEN 1 ELSE 0 END)
                ::NUMERIC / COUNT(*) * 100, 2)                         AS repeat_rate_pct,
            ROUND(AVG(order_count)::NUMERIC, 3)                        AS avg_orders_per_customer
        FROM counts
    """,
    "8_Freight_By_State": """
        SELECT
            c.customer_state,
            COUNT(DISTINCT o.order_id)                              AS total_orders,
            ROUND(AVG(oi.price)::NUMERIC, 2)                        AS avg_item_price,
            ROUND(AVG(oi.freight_value)::NUMERIC, 2)                AS avg_freight,
            ROUND((AVG(oi.freight_value / NULLIF(oi.price,0))
                * 100)::NUMERIC, 1)                                  AS freight_pct_of_price
        FROM olist.olist_order_items oi
        JOIN olist.olist_orders       o ON oi.order_id   = o.order_id
        JOIN olist.olist_customers    c ON o.customer_id = c.customer_id
        WHERE o.order_status NOT IN ('canceled','unavailable')
        GROUP BY c.customer_state
        ORDER BY freight_pct_of_price DESC
    """,
}

def style_sheet(ws):
    HEADER = PatternFill("solid", fgColor="1A3C5E")
    HFONT  = Font(color="FFFFFF", bold=True, size=11)
    ALT    = PatternFill("solid", fgColor="EEF4FB")
    BORDER = Border(
        bottom=Side(style="thin", color="CCCCCC"),
        right =Side(style="thin", color="EEEEEE"),
    )
    for cell in ws[1]:
        cell.fill = HEADER; cell.font = HFONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    for i, row in enumerate(ws.iter_rows(min_row=2), 2):
        for cell in row:
            if i % 2 == 0: cell.fill = ALT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 40)
    ws.freeze_panes = "A2"

def run():
    print("Connecting to PostgreSQL...")
    engine = get_engine()

    print(f"Exporting {len(QUERIES)} queries to Excel...")
    print(f"Output: {REPORT_FILE}\n")

    with pd.ExcelWriter(REPORT_FILE, engine="openpyxl") as writer:
        for name, sql in QUERIES.items():
            print(f"  Running {name}...", end=" ", flush=True)
            try:
                df = pd.read_sql_query(text(sql), engine.connect())
                df.to_excel(writer, sheet_name=name[:31], index=False)
                print(f"✅ {len(df):,} rows")
            except Exception as e:
                print(f"❌ {e}")

        # README sheet
        meta = pd.DataFrame({
            "Field": ["Project","Dataset","Generated","Queries"],
            "Value": ["Olist E-Commerce Analysis",
                      "Brazilian E-Commerce Public Dataset",
                      datetime.now().strftime("%Y-%m-%d %H:%M"),
                      str(len(QUERIES))]
        })
        meta.to_excel(writer, sheet_name="README", index=False)

    print("\nApplying formatting...")
    wb = load_workbook(REPORT_FILE)
    for sheet in wb.sheetnames:
        style_sheet(wb[sheet])
    wb.save(REPORT_FILE)

    print(f"\n✅ Done! File saved to:\n   {REPORT_FILE}")
    print("\nOpening folder...")
    os.system(f'open "{OUTPUT_DIR}"')

if __name__ == "__main__":
    run()
